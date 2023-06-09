import configparser
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
import random

config = configparser.ConfigParser(inline_comment_prefixes="#")
config.read('config.ini') 
Y_MIN = config.getint("graph_info", 'Y_MIN')
Y_MAX = config.getint("graph_info", 'Y_MAX')

def createGraph(list_data, trade_list_of_obj=[], infoList=[], yMin=Y_MIN, yMax=Y_MAX):

    file_name =f'analysis/{input("Введите имя создаваемого файла: ")}.xlsx'

    wb = Workbook()
    ws = wb.active
    for index, value in enumerate(list_data):
        ws.cell(index+1, 1, index)
        ws.cell(index+1, 2, value)

    for obj in trade_list_of_obj:
        index = obj['index']
        course = obj['course']
        operation = obj['operation']

        if operation == 'buy':
            ws.cell(index+1, 3, course)
        elif operation == 'sell':
            ws.cell(index+1, 4, course)
        else:
            ws.cell(index+1, 5, course)
        ws.cell(index+1, 6, obj.get('size', ''))
        ws.cell(index+1, 7, obj.get('rubles', ''))
        ws.cell(index+1, 8, obj.get('dollars', ''))

    list_data_length = len(list_data)

    chCourse = LineChart()
    yvalues = Reference(ws, min_col=2, min_row=1, max_row=list_data_length)
    chCourse.add_data(yvalues)
    yvalues = Reference(ws, min_col=3, min_row=1, max_row=list_data_length)
    chCourse.add_data(yvalues)
    yvalues = Reference(ws, min_col=4, min_row=1, max_row=list_data_length)
    chCourse.add_data(yvalues)
    yvalues = Reference(ws, min_col=5, min_row=1, max_row=list_data_length)
    chCourse.add_data(yvalues)

    chCourse.legend.position = "tr"
    chCourse.y_axis.scaling.min = yMin
    chCourse.y_axis.scaling.max = yMax

    ws.add_chart(chCourse, "L2")

    # Adding some information
    for index, info in enumerate(infoList):
        ws.cell(index+1, 9, info['title'])
        ws.cell(index+1, 10, info['content'])

    wb.save(file_name)


def createModel(initialCourse, endCourse):

    str_percent_range = input('Диапазон: ')
    fromValue, toValue = str_percent_range.split('=')
    fromValue, toValue = [float(fromValue), float(toValue)]
    possibility_rate = float(input('Соотношение положительных: '))

    current = initialCourse
    list_data = [current]

    while current < endCourse if endCourse > initialCourse else current > endCourse:
        plus = random.uniform(1, 100) <= possibility_rate
        difference = random.uniform(fromValue, toValue) / 100

        if plus:
            current += current * difference
        else:
            current -= current * difference

        list_data.append(round(current, 2))

    return list_data
