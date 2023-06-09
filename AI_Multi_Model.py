from AI import sort_out
from openpyxl import Workbook

print('Введите имена моделей:')
model_names = []
while True:
    model_name = input()
    if model_name:
        model_names.append(model_name)
    else:
        break

m_rs = []

for model_name in model_names:
    print(f'В процессе: {model_name}')
    sort_obj = sort_out(model_name, isWb=False)
    m_rs.append(sort_obj["data"])
    initialRubles = sort_obj["initialRubles"]
    print()

m_r_f = m_rs[0]

result = []
for p in m_r_f.keys():
    for r in m_r_f[p].keys():
        for s in m_r_f[p][r].keys():
            ok = True
            for m_r in m_rs[1:]:
                if not (m_r.get(p) and m_r[p].get(r) and m_r[p][r].get(s)):
                    ok = False
                    break

            if ok:
                money_sum = 0
                model_money_info = {}
                for index, m_r in enumerate(m_rs):
                    money = round(m_r[p][r][s]['resultRubles'], 2)
                    opCount = m_r[p][r][s]['opCount']
                    money_sum += money
                    model_money_info[model_names[index]] = {"money": money, 'opCount': opCount}

                result.append({'money_sum': money_sum, 'p': p, 'r': r, 's': s, 'model_money_info': model_money_info}) 

wb = Workbook()
ws = wb.active
ws.cell(1, 2, 'Начальный капитал')
ws.cell(1, 3, initialRubles)

ws.cell(3, 2, "Стратегия")
ws.cell(3, 3, "Процент")
ws.cell(3, 4, "Правило закупки")
col = 5
for model_name in model_names:
    ws.cell(3, col, model_name)
    col += 1

ws.cell(3, col, 'Доходность')

col += 1
for model_name in model_names:
    ws.cell(3, col, model_name)
    col += 1

row = 4
for strategy in result:
    ws.cell(row, 2, strategy["s"])
    ws.cell(row, 3, strategy["p"])
    ws.cell(row, 4, strategy["r"])
    col = 5
    for model_name in model_names:
        money = strategy["model_money_info"][model_name]["money"]
        ws.cell(row, col, money)
        col += 1
    
    ws.cell(row, col, strategy["money_sum"])

    col += 1
    for model_name in model_names:
        opCount = strategy["model_money_info"][model_name]["opCount"]
        ws.cell(row, col, opCount)
        col += 1

    row += 1

file_name = f'analysis/{input("Введите имя создаваемого файла: ")}.xlsx'
while True:
    try:
        print('Сохранение...')
        wb.save(file_name)
        print('Сохранено!')
        input('Нажмите любую клавишу для выхода...')
        break
    except:
        input("Не удалось сохранить файл. Закройте файл с тем же именем, позже нажмите enter для повторной попытки")

