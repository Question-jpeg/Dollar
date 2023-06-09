from dollarGraphMaker import createGraph
from robots import roboTrace, roboStas
from openpyxl import load_workbook


wb = load_workbook(
    f'analysis/{input("Введите имя файла excel с моделью: ")}.xlsx', data_only=True)
ws = wb.active

prices = [float(cell.value) for cell in ws['B']]

wb.close()

rubles = int(input("Введите начальную сумму рублей: "))
dollars = int(input("Введите начальную сумму долларов: "))
initialRubles = dollars * prices[0] + rubles

percentStep = float(input("Введите шаговый процент: "))
# robotName = input("Введите название робота (stas/trace): ")
robotName = 'stas'
buy_fix = ''
while buy_fix not in ['y', 'n']:
    if robotName == 'stas':
        buy_fix = input("Продавать только когда курс выше первой закупки? (y/n): ").lower()
    elif robotName == 'trace':
        buy_fix = input("Сделать условную покупку в начале графика? (y/n): ").lower()

buy_fix = buy_fix == 'y'
if robotName == 'stas':
    data = roboStas(prices, rubles, dollars, percentStep, buy_fix)
elif robotName == 'trace':
    data = roboTrace(prices, rubles, dollars, percentStep, buy_fix)



resultRubles = data['dollars'] * prices[len(prices)-1] + data['rubles']
operationsTrace = data['operationsTrace']

print(f'\n{data["count"]}')
print(initialRubles)
print(resultRubles)

createGraph(prices, operationsTrace, 
[{ "title": "Операций купля продажи:", 'content': data['count'] }, 
{ "title": "Изначальный капитал в рублях:", 'content': initialRubles }, 
{ "title": "Конечный капитал в рублях:", 'content': resultRubles },
{ "title": 'Комиссия:', 'content': data['commission'] }, 
{ "title": "", "content": "" },
{ "title": "Шаговый процент:", 'content': percentStep },
{ "title": "Распределение на покупку:", 'content': f'[{", ".join([str(item) for item in data["distribution"]])}]' }])

