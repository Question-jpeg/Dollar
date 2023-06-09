import configparser
from openpyxl import load_workbook, Workbook
from roboApi import roboStas
from AI_variations import VARIATIONS;
from reprint import output


def sort_out(model_name = "", isWb = True):
    if not model_name:
        model_name = input("Введите имя файла excel с моделью: ")

    config = configparser.ConfigParser(inline_comment_prefixes="#")
    config.read('config.ini') 

    RUBLES = config.getfloat('AI_info', 'RUBLES')
    DOLLARS = config.getfloat('AI_info', 'DOLLARS')
    FROM_PERCENT = config.getfloat('AI_info', 'FROM_PERCENT')
    TO_PERCENT = config.getfloat('AI_info', 'TO_PERCENT')
    PERCENT_RANGE_STEP = config.getfloat('AI_info', 'PERCENT_RANGE_STEP')
    PSR = [int(FROM_PERCENT*10), int(TO_PERCENT*10), int(PERCENT_RANGE_STEP*10)]

    wb = load_workbook(f'analysis/{model_name}.xlsx', data_only=True)
    ws = wb.active
    prices = [float(cell.value) for cell in ws['B']]
    wb.close()

    initialRubles = DOLLARS * prices[0] + RUBLES

    if isWb:    
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "Модель:")
        ws.cell(1, 2, model_name)
        ws.cell(1, 3, "Начальный капитал:")
        ws.cell(1, 4, initialRubles)

        ws.cell(3, 1, "Количество операций")
        ws.cell(3, 2, "Соотношение (Покупки:Продажи)")
        ws.cell(3, 3, "Шаговый процент")
        ws.cell(3, 4, "Правило первой закупки")
        ws.cell(3, 5, "Конечный капитал")
        ws.cell(3, 6, 'Комиссия')
        
        row = 4

    if not isWb:
        returnData = {}

    VARIATIONS_LENGTH = len(VARIATIONS)
    with output(output_type="list", initial_len=2) as output_list:
        for i in range(VARIATIONS_LENGTH):
            buyConfig = VARIATIONS[i]
            mainProgress = round(i / (VARIATIONS_LENGTH-1) * 100, 2)
            mainProgressBar = f'[{"#" * int(mainProgress // 5) + " " * (20 - int(mainProgress // 5))}]'

            for j in range(VARIATIONS_LENGTH):
                saleConfig = VARIATIONS[j]                
                secondaryProgress = round(j / (VARIATIONS_LENGTH-1) * 100, 2)
                secondaryProgressBar = f'[{"#" * int(secondaryProgress // 5) + " " * (20 - int(secondaryProgress // 5))}]'

                output_list[0] = f' {mainProgressBar} {mainProgress} %'
                output_list[1] = f' {secondaryProgressBar} {secondaryProgress} %'

                for percentStep in range(PSR[0], PSR[1] + PSR[2], PSR[2]):
                    percentStep /= 10
                    for buyFix in [True, False]:
                        data = roboStas(prices, RUBLES, DOLLARS, percentStep, buyConfig, saleConfig, buyFix)
                        resultRubles = data['dollars'] * prices[-1] + data['rubles']
                        count = data['count']
                        commission = round(data['commission'], 2)

                        buyConfigString = " ".join([str(n) for n in buyConfig])
                        saleConfigString = " ".join([str(n) for n in saleConfig])
                        strategy = f'[{buyConfigString}] [{saleConfigString}]'

                        if isWb:
                            ws.cell(row, 1, count)
                            ws.cell(row, 2, strategy)
                            ws.cell(row, 3, percentStep)
                            ws.cell(row, 4, buyFix)
                            ws.cell(row, 5, resultRubles)
                            ws.cell(row, 6, commission)

                            row += 1

                        if not isWb and (resultRubles > initialRubles):
                            returnData[percentStep] = returnData.get(percentStep, {})
                            returnData[percentStep][buyFix] = returnData[percentStep].get(buyFix, {})
                            returnData[percentStep][buyFix][strategy] = {"opCount": count, 'resultRubles': resultRubles}

    if isWb:
        file_name = f'analysis/{input("Введите имя создаваемого файла: ")}.xlsx'
        print('Сохранение...')
        wb.save(file_name)
        print('Сохранено!')
        input('Нажмите любую клавишу для выхода...')

    if not isWb:
        return {'data': returnData, 'initialRubles': initialRubles}

if __name__ == "__main__":
    sort_out()